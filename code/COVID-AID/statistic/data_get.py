# -*- coding: utf-8 -*-
"""
Created on Mon Aug 10 21:47:40 2020

@author: ljc545w
"""

import json
import openpyxl
import requests
from lxml import etree


def init():
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.116 Safari/537.36'
    }
    url = ' '
    response = requests.get(url=url, headers=headers)
    tree = etree.HTML(response.text)
    dict1 = tree.xpath('//script[@id="captain-config"]/text()')
    dict2 = json.loads(dict1[0])
    f = open('./statistic/covid_19.json', 'w', encoding='utf-8')
    f.write(str(dict2))
    f.close()
    return dict2


def china_total_data(data):
    wb = openpyxl.Workbook()
    ws_china = wb.active
    ws_china.title = '中国省份疫情数据'
    ws_china.append(['省/直辖市/自治区/行政区', '现有确诊', '累计确诊', '累计治愈',
                     '累计死亡', '现有确诊增量', '累计确诊增量',
                     '累计治愈增量', '累计死亡增量'])
    china = data['component'][0]['caseList']
    for province in china:
        ws_china.append([
            province['area'],
            province['curConfirm'],
            province['confirmed'],
            province['crued'],
            province['died'],
            province['curConfirmRelative'],
            province['confirmedRelative'],
            province['curedRelative'],
            province['diedRelative']
        ])
    ws_city = wb.create_sheet('中国城市疫情数据')
    ws_city.append([
        '城市', '现有确诊', '累计确诊', '累计治愈', '累计死亡', '累计确诊增量'
    ])
    for province in china:
        for city in province['subList']:
            if 'curConfirm' not in city:
                city['curConfirm'] = '0'
            if city['crued'] == '':
                city['crued'] = '0'
            if city['died'] == '':
                city['died'] = '0'
            if 'confirmedRelative' not in city:
                city['confirmedRelative'] = '0'
            ws_city.append([
                city['city'], '0', city['confirmed'],
                city['crued'], city['died'], city['confirmedRelative']
            ])
    time_domestic = data['component'][0]['mapLastUpdatedTime']
    ws_time = wb.create_sheet('中国疫情数据更新时间')
    ws_time.column_dimensions['A'].width = 22
    ws_time.append(['中国疫情数据更新时间'])
    ws_time.append([time_domestic])
    return china_daily_data(data, wb)


def global_total_data(data):
    wb = openpyxl.Workbook()
    ws_global = wb.active
    ws_global.title = '全球各国疫情数据'

    countries = data['component'][0]['caseOutsideList']
    ws_global.append(['国家', '现有确诊', '累计确诊', '累计治愈', '累计死亡', '累计确诊增量'])
    for country in countries:
        ws_global.append([
            country['area'],
            country['curConfirm'],
            country['confirmed'],
            country['crued'],
            country['died'],
            country['confirmedRelative']
        ])
    continent = data['component'][0]['globalList']
    for area in continent:
        ws_foreign = wb.create_sheet(area['area'] + '疫情数据')
        ws_foreign.append(['国家', '现有确诊', '累计确诊', '累计治愈', '累计死亡', '累计确诊增量'])
        for country in area['subList']:
            ws_foreign.append([
                country['country'],
                country['curConfirm'],
                country['confirmed'],
                country['crued'],
                country['died'],
                country['confirmedRelative']
            ])
    ws1, ws2 = wb['全球各国疫情数据'], wb['亚洲疫情数据']
    original_data = data['component'][0]['summaryDataIn']
    add_china_data = ['中国',
                      original_data['curConfirm'],
                      original_data['confirmed'],
                      original_data['cured'],
                      original_data['died'],
                      original_data['confirmedRelative']
                      ]
    ws1.append(add_china_data)
    ws2.append(add_china_data)
    time_foreign = data['component'][0]['foreignLastUpdatedTime']
    ws_time = wb.create_sheet('全球疫情数据更新时间')
    ws_time.column_dimensions['A'].width = 22
    ws_time.append(['全球疫情数据更新时间'])
    ws_time.append([time_foreign])

    return foreign_daily_data(data, wb)


def china_daily_data(data, wb):
    ccd_dict = data['component'][0]['trend']
    update_date = ccd_dict['updateDate']
    china_confirmed = ccd_dict['list'][0]['data']
    china_crued = ccd_dict['list'][2]['data']
    china_died = ccd_dict['list'][3]['data']
    china_surplus = []
    for i in range(len(update_date)):
        surplus = china_confirmed[i] - china_crued[i] - china_died[i]
        china_surplus.append(surplus)

    ws_china_surplus = wb.create_sheet('中国每日现有确诊数据')
    ws_china_surplus.append(['日期', '数据'])
    for data in zip(update_date, china_surplus):
        ws_china_surplus.append(data)

    ws_china_confirmed = wb.create_sheet('中国每日累计确诊数据')
    ws_china_confirmed.append(['日期', '数据'])
    for data in zip(update_date, china_confirmed):
        ws_china_confirmed.append(data)

    ws_china_crued = wb.create_sheet('中国每日累计治愈数据')
    ws_china_crued.append(['日期', '数据'])
    for data in zip(update_date, china_crued):
        ws_china_crued.append(data)

    ws_china_died = wb.create_sheet('中国每日累计死亡数据')
    ws_china_died.append(['日期', '数据'])
    for data in zip(update_date, china_died):
        ws_china_died.append(data)

    return wb


def foreign_daily_data(data, wb):
    te_dict = data['component'][0]['allForeignTrend']
    update_date = te_dict['updateDate']
    foreign_confirmed = te_dict['list'][0]['data']
    foreign_crued = te_dict['list'][1]['data']
    foreign_died = te_dict['list'][2]['data']
    foreign_surplus = []
    for i in range(len(update_date)):
        surplus = foreign_confirmed[i] - foreign_crued[i] - foreign_died[i]
        foreign_surplus.append(surplus)

    ws_foreign_surplus = wb.create_sheet('境外每日现有确诊数据')
    ws_foreign_surplus.append(['日期', '数据'])
    for data in zip(update_date, foreign_surplus):
        ws_foreign_surplus.append(data)

    ws_foreign_confirmed = wb.create_sheet('境外每日累计确诊数据')
    ws_foreign_confirmed.append(['日期', '数据'])
    for data in zip(update_date, foreign_confirmed):
        ws_foreign_confirmed.append(data)

    ws_foreign_crued = wb.create_sheet('境外每日累计治愈数据')
    ws_foreign_crued.append(['日期', '数据'])
    for data in zip(update_date, foreign_crued):
        ws_foreign_crued.append(data)

    ws_foreign_died = wb.create_sheet('境外每日累计死亡数据')
    ws_foreign_died.append(['日期', '数据'])
    for data in zip(update_date, foreign_died):
        ws_foreign_died.append(data)

    return wb
